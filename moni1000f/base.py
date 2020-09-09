import csv
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
from openpyxl import load_workbook


class MonitoringData(object):
    """
    Data class for working with ecosystem monitoring data.

    Originally designed for the data of the Monitoring Sites 1000 projects,
    but it can be used to handle similar data. Like the Pandas DataFrame object,
    the MonitoringData object contains the 'columns' and 'values' attributes.
    The original two-dimensional array is contained as 'data'.

    Parameters
    ----------
    data : numpy ndarray
        Two-dimensional numpy ndarray with the dtype of unicode string ('U')
    header : bool, default True
        If the input data includes a header line
    plot_id : str, optional
        Plot ID
    data_type : str, optional
        Data type. It will be guessed from header names if no given
    comments : numpy ndarray, optional
        Two-dimensional numpy ndarray of comment lines (rows) of data
    metadata : dict, optional
        Metadata for the input data

    Attributes
    ----------
    data : numpy ndarray
        Original Data.
    values : numpy ndarray
        Data values
    columns : list
        List of column names
    plot_id : str
        Plot ID
    data_type : str
        Data type. Typical values are: tree, litter, seed or other
    comments : numpy ndarray
        Comment lines
    data_with_comments : numpy ndarray
        Data with comment lines
    metadata : dict
        Metadata

    """
    def __init__(self,
                 data: np.ndarray = np.array([]),
                 header: bool = True,
                 plot_id: Optional[str] = None,
                 data_type: Optional[str] = None,
                 metadata: Dict[str, str] = {},
                 comments: Optional[np.ndarray] = None,
                 *args,
                 **kwargs):

        self.data = data
        self.plot_id = plot_id
        if len(data) > 1:
            self.header = header
        else:
            self.header = False
        if data_type:
            self.data_type = data_type
        else:
            self.data_type = self.__guess_data_type()
        self.metadata = metadata
        self.comments = comments

    @property
    def values(self):
        if self.header:
            return self.data[1:]
        else:
            return self.data

    @property
    def columns(self):
        if self.header:
            return self.data[0]
        else:
            return np.array([])

    @property
    def data_with_comments(self):
        return join_comments(self.data, self.comments)

    def __repr__(self):
        s1 = "data_shape={}".format(self.values.shape)
        s2 = "plot_id={!r}".format(self.plot_id) if self.plot_id else ""
        s3 = "data_type={!r}".format(self.data_type) if self.data_type else ""
        s = ", ".join([i for i in (s1, s2, s3) if i])
        return "{}({})".format(self.__class__.__name__, s)

    def __getitem__(self, key):
        if isinstance(key, str):
            data_s = self.select_cols(key, add_header=True)
        elif isinstance(key, list) and all([isinstance(i, str) for i in key]):
            data_s = self.select_cols(key, add_header=True)
        elif isinstance(key, slice):
            data_s = np.vstack((self.columns, self.values[key]))
        elif isinstance(key, tuple) and all([isinstance(i, slice) for i in key]):
            data_s = np.vstack((self.columns[key[1]], self.values[key]))
        elif isinstance(key, tuple) and isinstance(key[1], slice):
            data_s = np.vstack((self.columns[key[1]], self.values[key]))
        elif isinstance(key, tuple) and isinstance(key[0], slice):
            data_s = np.append(self.columns[key[1]], self.values[key])
        elif isinstance(key, int):
            data_s = np.vstack((self.columns, self.values[key]))
        elif isinstance(key, np.ndarray) and key.dtype == "bool":
            if len(key.shape) == 1:
                data_s = np.vstack((self.columns, self.values[key]))
            else:
                return self.values[key]
        else:
            return self.values[key]
        return self.__getitem_return(data_s,
                                     header=self.header,
                                     plot_id=self.plot_id,
                                     data_type=self.data_type,
                                     metadata=self.metadata,
                                     comments=self.comments)

    @classmethod
    def __getitem_return(cls, data, **kwargs):
        return cls(data, **kwargs)

    def __guess_data_type(self):
        """Guess data type from the header line of the data."""
        cols_t = [
            "tag_no",
            "indv_no",
            "spc_japan",
            "^gbh[0-9]{2}$",
            "^s_date[0-9]{2}$",
        ]

        cols_l = [
            "^trap_id$",
            "^s_date1$",
            "^s_date2$",
            "^wdry_",
            "^w_",
        ]

        cols_s = [
            "^trap_id$",
            "^s_date1$",
            "^s_date2$",
            "^w",
            "^spc$",
            "^status$",
            "^form$",
        ]

        if all([any(filter(lambda x: re.match(i, x), self.columns)) for i in cols_t]):
            data_type = "tree"
        elif all([any(filter(lambda x: re.match(i, x), self.columns)) for i in cols_l]):
            data_type = "litter"
        elif all([any(filter(lambda x: re.match(i, x), self.columns)) for i in cols_s]):
            data_type = "seed"
        else:
            data_type = "other"

        return data_type

    def select_cols(self,
                    col: Union[str, List[str], None] = None,
                    regex: Union[str, None] = None,
                    add_header: bool = False) -> np.ndarray:
        """
        Select data columns.

        Paramters
        ---------
        col : str, list, optional
            Column name(s) to select
        regex : str
            Regular expression pattern. Columns will be selected if the column names
            matches with the regular expression pattern
        add_header : bool, default False
            Add a header row to an output array

        """
        start = 1 if (self.header and not add_header) else 0

        if not self.header:
            raise RuntimeError("The data does not have column names")

        if regex:
            r = re.compile(regex)
            match = np.vectorize(lambda x: True if r.match(x) else False)(self.columns)
            selected = self.data[start:, match]

        elif col:
            if isinstance(col, str):
                col = [col]
            isin = np.isin(col, self.columns)
            if all(isin):
                selected = self.data[start:, np.isin(self.columns, col)]
            else:
                notin = ", ".join(np.array(col)[~isin])
                s = "s" if (~isin).sum() > 1 else ""
                msg = "Undefined column{}: {}".format(s, notin)
                raise KeyError(msg)

        else:
            raise RuntimeError("col or regex is needed")

        if not add_header and selected.shape[1] == 1:
            selected = selected.flatten()

        return selected

    def to_csv(self,
               outpath: str,
               keep_comments: bool = True,
               cleaning: bool = True,
               encoding: str = "utf-8",
               na_rep: str = ""):
        """
        Export data as a csv file.

        Parameters
        ----------
        outpath : str
            Path to the output file
        keep_comments : bool, default True
            If include the comment lines
        cleaning : bool, default True
            If clean up the data
        encoding : str, default "utf-8"
            Text encoding
        na_rep : str, default ""
            Replacement characters for NaN

        """
        if keep_comments:
            data = self.data_with_comments
        else:
            data = self.data

        data_to_csv(data,
                    outpath=outpath,
                    cleaning=cleaning,
                    encoding=encoding,
                    na_rep=na_rep)


def read_file(filepath: Union[str, Path]) -> np.ndarray:
    """
    Read a data file (.csv or .xlsx) and return a np.ndarray object.

    データファイルの読み込み。全て文字列として読み込む。

    Parameters
    ----------
    filepath : str or pathlibs Path object
        Path to the data file

    """
    filepath = Path(filepath).expanduser()
    suffix = filepath.suffix
    if suffix == ".xlsx":
        wb = load_workbook(filepath, read_only=True, data_only=True)
        if "Data" in wb.sheetnames:
            ws = wb["Data"]
        else:
            ws = wb[wb.sheetnames[0]]
        data = np.array([[str(j) if j is not None else "" for j in i]
                         for i in ws.values])
        wb.close()
    elif suffix == ".csv":
        with filepath.open() as f:
            reader = csv.reader(f)
            data = np.array([i for i in reader])
    else:
        msg = ("{} file is not supported. Supported formats are: "
               ".csv, .xlsx, .xlsm, .xltx, .xltm").format(suffix)
        raise RuntimeError(msg)

    # remove blank rows/columns
    data = mat_strip(data, strip="")

    return data


def mat_strip(mat: np.ndarray, strip: Any = "") -> Optional[np.ndarray]:
    """
    Strip rows and columns from a tow-dimentional array.

    Parameters
    ----------
    mat : numpy ndarray
        Two-dimentional array
    strip : Any, default ""
        If all elements have this value, remove the rows/columns from the edges.

    """
    match = np.vectorize(lambda x: True if x == strip else False)(mat)
    if not match.any():
        return mat
    elif match.all():
        return np.array([], mat.dtype)

    x = match.all(axis=0).tolist()
    y = match.all(axis=1).tolist()

    def strip_index(z: List[bool], right: bool = False):
        z = z.copy()
        pop_i = -1 if right else 1
        i = 0
        while True:
            if not z.pop(pop_i):
                break
            i += 1
        return i

    mat = mat[:, slice(strip_index(x), len(x) - strip_index(x, True))]
    mat = mat[slice(strip_index(y), len(y) - strip_index(y, True)), :]

    return mat


def split_comments(data: np.ndarray,
                   comment_chr: str = "#") -> Tuple[np.ndarray, np.ndarray]:
    """Split comment lines (rows) from data array."""
    if comment_chr:
        com_rows = np.vectorize(lambda x: str(x).startswith(comment_chr))(data[:, 0])
        comments = data[com_rows]
        data = data[~com_rows]
    else:
        comments = np.ndarray(shape=(0, data.shape[1]), dtype=data.dtype)

    if comments.size > 0:
        comments = mat_strip(comments, strip="")

    return data, comments


def join_comments(data: np.ndarray, comments: np.ndarray) -> np.ndarray:
    """Join comment lines to data array."""
    if comments.size == 0:
        return data

    if (comments.shape[1] - data.shape[1]) < 0:
        spacer = np.full((comments.shape[0], data.shape[1] - comments.shape[1]), "")
        comments_new = np.hstack((comments, spacer))
        data_new = np.vstack((comments_new, data))
    elif (comments.shape[1] - data.shape[1]) > 0:
        spacer = np.full((data.shape[0], comments.shape[1] - data.shape[1]), "")
        data_new = np.hstack((data, spacer))
        data_new = np.vstack((comments, data_new))
    else:
        data_new = np.vstack((comments, data))

    return data_new


def get_metadata(comments: np.ndarray) -> Dict[str, str]:
    """Get metadata from comment lines."""
    keys = [
        'DATA CREATED', 'DATA CREATER', 'DATA TITLE', 'SITE NAME', 'PLOT NAME',
        'PLOT ID', 'PLOT SIZE', 'NO. OF TRAPS', 'TRAP SIZE'
    ]
    metadata = {}
    for key in keys:
        match = list(zip(*np.where(comments == key)))
        if match:
            i, j = match[0]
            metadata[key] = comments[i, j + 2]
    return metadata


def get_plotid(filepath: Union[str, Path]) -> str:
    """
    Get the plot id from a file name.

    Parameters
    ----------
    filepath : str or pathlibs Path object
        Path to the data file

    """
    filepath = Path(filepath)
    ftype = ["AT", "EC", "BC", "EB", "DB"]
    r = re.compile("|".join(["[A-Z]{{2}}-{}[0-9]".format(i) for i in ftype]))
    m = r.search(filepath.name)
    return m.group() if m else ""


def read_data(
    filepath: str,
    comment_chr: str = "#",
    header: bool = True,
    plot_id: Optional[str] = None,
    data_type: Optional[str] = None,
    metadata: Optional[Dict[str, str]] = None,
) -> MonitoringData:
    """
    Read a data file and return a MonitoringData object.

    Parameters
    ----------
    filepath: str
        Path to the data file. Supported formats are: .csv, .xlsx, .xlsm, .xltx, .xltm
    comment_chr: str, default "#"
        Character to detect commented lines. If found at the beginning of a line, the
        line will be parsed as commented lines and split from remaininig lines
    header: bool, default True
        If the parsed data includes a header line
    plot_id : str, optional
        Plot ID
    data_type : str, optional
        Data type. It will be guessed from header names if no given
    metadata : dict, optional
        Metadata for the input data

    """
    data = read_file(filepath)
    data, comments = split_comments(data, comment_chr)

    if not metadata:
        if len(comments) > 0:
            metadata = get_metadata(comments)
        else:
            metadata = {}

    if not plot_id:
        if "PLOT ID" in metadata:
            plot_id = metadata["PLOT ID"]
        else:
            plot_id = get_plotid(filepath)

    return MonitoringData(data,
                          plot_id=plot_id,
                          data_type=data_type,
                          metadata=metadata,
                          header=header,
                          comments=comments)


def clean_data(data: np.ndarray) -> np.ndarray:
    """
    Clean up the data.

    Remove white spaces, line breaks, normalize unocode characters etc.).

    Parameters
    ----------
    data: numpy ndarray
        Two dimentional array with the dtype of string('<U')

    """
    # 不要な空白を削除
    data = np.vectorize(lambda x: x.strip())(data)
    # unicode文字列の標準化（全角英数字->半角英数字, 半角カナ->全角カナ, etc.）
    data = np.vectorize(lambda x: unicodedata.normalize("NFKC", x))(data)
    # float型で小数点以下が長くなっている値を丸める
    data = np.vectorize(clean_float)(data)
    # datetime型の文字列をyyyymmdd形式の文字列に変換
    data = np.vectorize(datetime_to_yyyymmdd)(data)
    # セル内改行・タブ・垂直タブ・改頁を削除
    data = np.vectorize(lambda x: re.sub("\r\n|\n|\r|\t|\x0b|\x0c", "", x))(data)

    return data


def clean_float(x: str, precision: str = "single") -> str:
    """
    Floating-point rounding for a numerical string.

    文字列が数値で、小数点以下が長い場合は丸める。

    Parameters
    ----------
    x : str
        Any string
    precision : str, default 'single'
        Precision for floating-point: 'single' or 'double'

    """
    try:
        int(x)
    except ValueError:
        try:
            if precision == "single":
                return str(np.float32(x))
            elif precision == "double":
                return str(np.float64(x))
            else:
                raise ValueError("Valid presision values are: single, double")
        except ValueError:
            return str(x)
    else:
        return str(x)


def datetime_to_yyyymmdd(s: str) -> str:
    """
    Convert a datetime string to a string in the yyyymmdd format.

    日付がyyyymmddになっていない場合修正。

    """
    pat_datetime = re.compile(r"(\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2})")
    m = pat_datetime.match(str(s))
    if m:
        dt = datetime.strptime(m.group(), "%Y-%m-%d %H:%M:%S")
        yyyymmdd = datetime.strftime(dt, "%Y%m%d")
        return yyyymmdd
    else:
        return s


def data_to_csv(data: np.ndarray,
                outpath: str,
                cleaning: bool = True,
                encoding: str = "utf-8",
                na_rep: str = ""):
    """
    Export a two-dementional numpy array as a csv file after cleaning.

    2次元配列データをクリーニング後、CSVファイルとして書き出す。

    Parameters
    ----------
    data : numpy array
        Two dimentional numpy array
    outpath : str
        Path to the output file
    cleaning : bool, default True
        If clean up the data
    encoding : str, default "utf-8"
        Text encoding
    na_rep : str, default ""
        Replacement characters for NaN

    """
    if cleaning:
        data = clean_data(data)

    data = np.where(data == "nan", na_rep, data)

    with open(outpath, "w", encoding=encoding) as f:
        writer = csv.writer(f, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        for row in data:
            writer.writerow(row)
